"""
Build BNF-chapter -> emission factor mapping for NHS medicines procurement.

VERSION 2 — All factors must trace to a citable source document.
See data/interim/defra_mapping_SOURCES.md for full citations.

═══════════════════════════════════════════════════════════════════════════════
WHAT WAS WRONG WITH VERSION 1
═══════════════════════════════════════════════════════════════════════════════

Version 1 contained chapter-differentiated kgCO2e/£ figures (ranging 0.58–1.15)
generated from general knowledge. Those numbers are NOT citable and have been
completely replaced below.

═══════════════════════════════════════════════════════════════════════════════
WHAT WE SEARCHED FOR AND FOUND
═══════════════════════════════════════════════════════════════════════════════

[FOUND — downloaded and used]
ONS (2026). "UK Environmental Accounts: Atmospheric emissions: greenhouse gas
emissions intensity by industry." Dataset 04atmosphericemissionsghgintensity.xlsx,
published 5 June 2026, covering 1990–2024. DOI/URL:
https://www.ons.gov.uk/economy/environmentalaccounts/datasets/
ukenvironmentalaccountsatmosphericemissionsgreenhousegasemissionsintensitybyeconomicsectorunitedkingdom

This dataset gives GHG emission intensity by SIC industry code in units of
THOUSAND TONNES CO2e PER GBP MILLION, which equals KG CO2e PER GBP.
It is the most granular publicly available UK source for sector-level emission
intensity. File saved: data/raw/ons_ghg_intensity_by_industry.xlsx

[FOUND — methodology only, no downloadable multiplier table]
NHS England (2022). "Delivering a 'Net Zero' National Health Service" (July 2022).
URL: https://www.england.nhs.uk/greenernhs/wp-content/uploads/sites/51/2022/07/
B1728-delivering-a-net-zero-nhs-july-2022.pdf
Pages 56–58 confirm supply chain emissions use a UK Multi-Region Input-Output
(MRIO) model (University of Leeds for Defra), using EEIO intensities for 105
economic sectors. The actual sector multiplier table is NOT published.
File saved: data/raw/nhs_delivering_net_zero_2022.pdf

[NOT FOUND — paywalled]
Tennison I et al. (2021). "Health care's response to climate change: a carbon
footprint assessment of the NHS in England." Lancet Planetary Health 5(2):
e84–e92. DOI: 10.1016/S2542-5196(21)00005-X
Full text returns HTTP 403. Supplementary data not accessible.

[NOT FOUND — URL does not exist]
Defra EEIO model tables. Multiple gov.uk URL patterns tried; all return 404.
The MRIO multiplier table referenced as "footnote 47" in the NHS Net Zero report
is not publicly downloadable.

[NOT FOUND — behind authentication]
DEFRA 2025 GHG Conversion Factors (already downloaded) — contains ZERO
spend-based (kgCO2e/£) factors. Confirmed from flat file inspection.

═══════════════════════════════════════════════════════════════════════════════
CRITICAL METHODOLOGICAL DISTINCTION
═══════════════════════════════════════════════════════════════════════════════

The ONS intensity data (what we have) = UK PRODUCTION-BASIS emissions.
This captures only the direct and indirect GHG from UK producers in SIC 21,
per £ of their gross value added. It EXCLUDES:
  - Emissions from imported active pharmaceutical ingredients and intermediates
  - Full supply chain upstream effects modelled in an MRIO
  - The global production of imported medicines (most branded/specialty drugs)

The NHS Carbon Footprint model uses CONSUMPTION-BASIS MRIO factors, which are
higher because they include imports' embedded emissions. The NHS report states
supply chain is ~60% of total NHS carbon footprint (~25 Mt total => ~15 Mt
supply chain), implying a much higher effective intensity than the ONS figure.

For the 2024 ONS data:
  SIC 21 (all pharma): 0.17 kgCO2e/£  — UK PRODUCTION-BASIS ONLY
  SIC C  (all manufacturing): 0.54 kgCO2e/£
  SIC Q  (human health): 0.02 kgCO2e/£

The production-basis SIC 21 figure (0.17) is the ONLY citable public figure.
It is a confirmed lower bound. The true consumption-basis figure used in NHS
Carbon Footprint modelling is higher, but those specific multipliers are not in
any publicly downloadable document.

All 16 BNF chapter entries therefore use SIC 21 = 0.17 kgCO2e/£ as the single
most defensible citable figure, with an explicit caveat that it is a lower bound.

═══════════════════════════════════════════════════════════════════════════════
"""

import os
import re
import pandas as pd
import openpyxl

RAW_DIR   = os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw")
INT_DIR   = os.path.join(os.path.dirname(__file__), "..", "..", "data", "interim")
DEFRA_FULL = os.path.join(RAW_DIR, "ghg_conversion_factors_2025_full_set.xlsx")
DEFRA_FLAT = os.path.join(RAW_DIR, "ghg_conversion_factors_2025_flat.xlsx")
ONS_INTENSITY = os.path.join(RAW_DIR, "ons_ghg_intensity_by_industry.xlsx")


# ─── Read the real ONS SIC 21 value ──────────────────────────────────────────

def get_ons_sic21_intensity(year=2024):
    """
    Extract SIC 21 GHG intensity from ONS dataset for a given year.
    Units: kgCO2e per GBP (thousand tonnes CO2e / GBP million).
    Column index 55 (1-based) = SIC 21 in the GHG intensity sheet.
    """
    wb = openpyxl.load_workbook(ONS_INTENSITY, read_only=True)
    ws = wb["GHG intensity"]
    rows = list(ws.iter_rows(values_only=True))
    col_0idx = 55 - 1  # convert 1-based to 0-based
    for row in rows[8:]:
        if row[0] == year:
            val = row[col_0idx]
            if isinstance(val, (int, float)):
                return float(val)
    raise ValueError(f"Year {year} not found in ONS intensity dataset")


def get_ons_all_intensities(year=2024):
    """Return dict of {SIC_label: intensity} for all target sectors."""
    wb = openpyxl.load_workbook(ONS_INTENSITY, read_only=True)
    ws = wb["GHG intensity"]
    rows = list(ws.iter_rows(values_only=True))
    sic_row = rows[6]
    name_row = rows[7]

    targets = {
        55: "SIC 21 — Basic pharmaceutical products and preparations",
        3:  "SIC C — Total manufacturing",
        17: "SIC Q — Human health and social work",
        123:"SIC 86 — Human health services",
        71: "SIC 32 — Other manufactured goods (incl. medical instruments)",
        88: "SIC 49.3-5 — Land transport excl. rail",
        91: "SIC 52 — Warehousing and support transport services",
        85: "SIC 46 — Wholesale trade excl. motor vehicles",
        134:"SIC Total — Total economy excl. consumer expenditure",
    }
    result = {}
    for row in rows[8:]:
        if row[0] == year:
            for col_1idx, label in targets.items():
                v = row[col_1idx - 1]
                result[label] = v if isinstance(v, (int, float)) else None
            break
    return result


# ─── BNF chapter definitions (same as v1) ─────────────────────────────────────

BNF_CHAPTERS = [
    (1, "Gastro-intestinal",
     r"antacid|omeprazol|lansoprazol|pantoprazol|ranitidine|metoclopram|domperidon|"
     r"loperamide|mesalazine|infliximab|adalimumab.*crohn|azathioprine|"
     r"lactulose|senna|bisacodyl|macrogol|linaclotide|prucalopride|"
     r"pancreatin|ursodeoxycholic|ondansetron|granisetron|aprepitant|"
     r"laxative|antidiarrhoeal|antispasmodic|hyoscine butylbromide|"
     r"sucralfate|colestyramine|obeticholic"),

    (2, "Cardiovascular",
     r"aspirin.*card|clopidogrel|warfarin|apixaban|rivaroxaban|dabigatran|edoxaban|"
     r"atorvastatin|rosuvastatin|simvastatin|pravastatin|ezetimibe|"
     r"amlodipine|nifedipine|diltiazem|verapamil|"
     r"bisoprolol|carvedilol|metoprolol|atenolol|"
     r"lisinopril|ramipril|perindopril|enalapril|candesartan|losartan|"
     r"furosemide|bumetanide|spironolactone|eplerenone|"
     r"digoxin|amiodarone|flecainide|"
     r"nicorandil|isosorbide|glyceryl trinitrate|"
     r"sacubitril|ivabradine|dapagliflozin.*heart|"
     r"antihypertensive|antiarrhythmic|anticoagulant|antiplatelet|statin"),

    (3, "Respiratory",
     r"salbutamol|salmeterol|formoterol|indacaterol|olodaterol|"
     r"budesonide|fluticasone|beclometasone|ciclesonide|mometasone.*inhal|"
     r"tiotropium|ipratropium|glycopyrronium|aclidinium|umeclidinium|"
     r"montelukast|zafirlukast|theophylline|aminophylline|"
     r"roflumilast|benralizumab|mepolizumab|dupilumab|"
     r"carbocisteine|erdosteine|dornase alfa|"
     r"inhaler|nebuliser solution|"
     r"bronchodilator|corticosteroid.*inhal"),

    (4, "Central Nervous System",
     r"sertraline|fluoxetine|citalopram|escitalopram|venlafaxine|duloxetine|"
     r"mirtazapine|amitriptyline|nortriptyline|clomipramine|"
     r"olanzapine|quetiapine|risperidone|aripiprazole|clozapine|"
     r"diazepam|lorazepam|clonazepam|midazolam|temazepam|"
     r"zopiclone|zolpidem|melatonin|"
     r"levodopa|co-careldopa|pramipexole|ropinirole|rasagiline|selegiline|"
     r"donepezil|rivastigmine|galantamine|memantine|"
     r"gabapentin|pregabalin|lamotrigine|levetiracetam|valproate|"
     r"carbamazepine|phenytoin|topiramate|"
     r"methylphenidate|atomoxetine|lisdexamfetamine|"
     r"sumatriptan|rizatriptan|zolmitriptan|"
     r"naltrexone|buprenorphine.*opioid|methadone|"
     r"antidepressant|antipsychotic|anxiolytic|sedative|hypnotic|anticonvulsant|"
     r"antiparkinsonian"),

    (5, "Infections",
     r"amoxicillin|co-amoxiclav|ampicillin|flucloxacillin|piperacillin|"
     r"cefalexin|cefuroxime|ceftriaxone|cefotaxime|ceftazidime|cefepime|"
     r"clarithromycin|azithromycin|erythromycin|"
     r"ciprofloxacin|levofloxacin|moxifloxacin|"
     r"trimethoprim|co-trimoxazole|nitrofurantoin|"
     r"metronidazole|tinidazole|"
     r"doxycycline|tetracycline|"
     r"vancomycin|teicoplanin|linezolid|daptomycin|"
     r"meropenem|ertapenem|imipenem|"
     r"fluconazole|itraconazole|voriconazole|posaconazole|caspofungin|"
     r"aciclovir|valaciclovir|ganciclovir|valganciclovir|"
     r"oseltamivir|remdesivir|"
     r"isoniazid|rifampicin|ethambutol|pyrazinamide|"
     r"antibiotic|antifungal|antiviral|antimicrobial|bactericidal"),

    (6, "Endocrine",
     r"insulin|metformin|glipizide|gliclazide|glibenclamide|"
     r"sitagliptin|vildagliptin|saxagliptin|alogliptin|"
     r"dapagliflozin|empagliflozin|canagliflozin|"
     r"liraglutide|semaglutide|dulaglutide|exenatide|"
     r"levothyroxine|liothyronine|propylthiouracil|carbimazole|"
     r"prednisolone|dexamethasone|hydrocortisone|fludrocortisone|"
     r"growth hormone|somatropin|"
     r"testosterone|anastrozole|letrozole|"
     r"antidiabetic|hypoglycaemic|thyroid|corticosteroid(?!.*inhal)"),

    (7, "Obstetrics, Gynaecology & Urinary",
     r"oxybutynin|solifenacin|tolterodine|mirabegron|"
     r"tamsulosin|alfuzosin|silodosin|doxazosin.*bph|finasteride|dutasteride|"
     r"oestrogen|estrogen|progesterone|norethisterone|"
     r"mifepristone|misoprostol|"
     r"sildenafil|tadalafil|vardenafil|avanafil|"
     r"oxytocin|ergometrine|atosiban|"
     r"clomifene|gonadotrophin|"
     r"contraceptive|intrauterine"),

    (8, "Malignant Disease & Immunosuppression",
     r"methotrexate|cyclophosphamide|chlorambucil|melphalan|"
     r"fluorouracil|capecitabine|gemcitabine|cytarabine|"
     r"doxorubicin|epirubicin|idarubicin|"
     r"paclitaxel|docetaxel|nab-paclitaxel|cabazitaxel|"
     r"carboplatin|cisplatin|oxaliplatin|"
     r"imatinib|dasatinib|nilotinib|bosutinib|ponatinib|"
     r"erlotinib|gefitinib|afatinib|osimertinib|"
     r"trastuzumab|pertuzumab|cetuximab|bevacizumab|"
     r"pembrolizumab|nivolumab|atezolizumab|durvalumab|"
     r"rituximab|obinutuzumab|ofatumumab|"
     r"bortezomib|carfilzomib|ixazomib|daratumumab|"
     r"lenalidomide|thalidomide|pomalidomide|"
     r"letrozole|anastrozole|exemestane|tamoxifen|fulvestrant|"
     r"leuprorelin|goserelin|triptorelin|"
     r"mycophenolate|tacrolimus|ciclosporin|sirolimus|"
     r"antineoplastic|chemotherapy|cytotoxic|immunosuppressant|"
     r"anticancer|monoclonal antibody|targeted therapy"),

    (9, "Nutrition & Blood",
     r"ferrous|ferric|iron.*inj|"
     r"vitamin b12|hydroxocobalamin|cyanocobalamin|folic acid|"
     r"epoetin|darbepoetin|erythropoietin|"
     r"filgrastim|pegfilgrastim|lenograstim|"
     r"albumin|"
     r"parenteral nutrition|total parenteral|"
     r"potassium chloride|sodium chloride.*inj|glucose.*inj|hartmann|"
     r"magnesium sulfate.*inj|calcium gluconate.*inj|"
     r"factor viii|factor ix|factor vii|"
     r"fresh frozen plasma|prothrombin|"
     r"alteplase|tenecteplase|streptokinase|"
     r"heparin|enoxaparin|dalteparin|fondaparinux|"
     r"antifibrinolytic|tranexamic|"
     r"anaemia|haemostatic|antianaemic"),

    (10, "Musculoskeletal & Joint Diseases",
     r"ibuprofen|naproxen|diclofenac|celecoxib|etoricoxib|"
     r"paracetamol(?!.*codeine).*oral|"
     r"methotrexate.*rheuma|sulfasalazine|hydroxychloroquine|"
     r"adalimumab|etanercept|certolizumab|golimumab|"
     r"tocilizumab|sarilumab|baricitinib|tofacitinib|upadacitinib|"
     r"abatacept|"
     r"alendronic|risedronate|zoledronic|ibandronic|"
     r"denosumab|teriparatide|romosozumab|"
     r"colchicine|allopurinol|febuxostat|"
     r"nsaid|dmard|disease modifying|antirheumatic"),

    (11, "Eye",
     r"latanoprost|bimatoprost|travoprost|timolol.*eye|"
     r"brimonidine|brinzolamide|dorzolamide|"
     r"ranibizumab|bevacizumab.*eye|aflibercept|faricimab|"
     r"dexamethasone.*eye|prednisolone.*eye|"
     r"chloramphenicol.*eye|fusidic acid.*eye|"
     r"eye drop|eye ointment|intravitreal"),

    (12, "Ear, Nose & Oropharynx",
     r"fluticasone.*nasal|mometasone.*nasal|beclometasone.*nasal|"
     r"azelastine.*nasal|ipratropium.*nasal|"
     r"ear drop|nasal spray|"
     r"nystatin.*oral|miconazole.*oral gel"),

    (13, "Skin",
     r"betamethasone.*cream|clobetasol|hydrocortisone.*cream|"
     r"tacrolimus.*skin|pimecrolimus|"
     r"aciclovir.*cream|fusidic acid.*cream|mupirocin|"
     r"calcipotriol|dithranol|coal tar|"
     r"isotretinoin|adapalene|benzoyl peroxide|"
     r"emollient|aqueous cream|"
     r"topical.*antibiotic|topical.*antifungal|"
     r"clotrimazole|miconazole.*skin|terbinafine.*skin"),

    (14, "Immunological Products & Vaccines",
     r"vaccine|vaccination|immunisation|"
     r"immunoglobulin|antitetanus|antitoxin|"
     r"infliximab|adalimumab(?!.*crohn)(?!.*rheuma)|"
     r"omalizumab|dupilumab(?!.*asthma)|"
     r"normal immunoglobulin|subcutaneous immunoglobulin|iv immunoglobulin"),

    (15, "Anaesthesia",
     r"propofol|thiopental|ketamine|etomidate|"
     r"suxamethonium|rocuronium|vecuronium|atracurium|mivacurium|"
     r"neostigmine|sugammadex|"
     r"fentanyl|alfentanil|remifentanil|sufentanil|"
     r"morphine|oxycodone|hydromorphone|"
     r"bupivacaine|levobupivacaine|ropivacaine|lidocaine|"
     r"sevoflurane|isoflurane|desflurane|nitrous oxide|"
     r"dexmedetomidine|clonidine.*anaes|"
     r"ondansetron.*anaes|cyclizine|dexamethasone.*anaes|"
     r"anaesthetic|analgesic.*inj|opioid.*inj|neuromuscular block"),

    (0, "Other / Unclassified", r""),
]


def classify_bnf_chapter(product_name):
    name = str(product_name).lower()
    for ch_num, ch_name, pattern in BNF_CHAPTERS:
        if ch_num == 0:
            return ch_num, ch_name
        if pattern and re.search(pattern, name):
            return ch_num, ch_name
    return 0, "Other / Unclassified"


# ─── Build mapping with REAL ONS data ─────────────────────────────────────────

def build_mapping_df(sic21_intensity):
    """
    All 16 BNF chapters receive the same ONS SIC 21 intensity (2024).
    No chapter-level differentiation is possible from public data sources.
    Confidence is set to the appropriate level per the source assessment below.
    """
    SOURCE = (
        "ONS (2026). UK Environmental Accounts: Atmospheric emissions: "
        "GHG emissions intensity by industry, 2024. "
        "Table: GHG intensity, SIC 21 (col 55). "
        "File: 04atmosphericemissionsghgintensity.xlsx. "
        "Published 5 June 2026. "
        "URL: https://www.ons.gov.uk/economy/environmentalaccounts/datasets/"
        "ukenvironmentalaccountsatmosphericemissionsgreenhousegasemissionsintensity"
        "byeconomicsectorunitedkingdom"
    )

    RATIONALE = (
        "UK production-basis emission intensity for SIC 21 (Basic pharmaceutical "
        "products and preparations), 2024. LOWER BOUND — excludes imported supply "
        "chain emissions. NHS Carbon Footprint model uses consumption-basis MRIO "
        "factors (higher) but those are not publicly downloadable. No public source "
        "exists to differentiate emission intensity by BNF chapter; a single SIC 21 "
        "factor is applied to all chapters. For research use, flag this as the "
        "production-basis floor; the true Scope 3 Category 1 intensity is higher."
    )

    rows = []
    for ch_num, ch_name, _ in BNF_CHAPTERS:
        rows.append({
            "bnf_chapter_num": ch_num,
            "bnf_chapter_name": ch_name,
            "sic_code": "SIC 21",
            "sic_description": "Basic pharmaceutical products and preparations",
            "intensity_basis": "UK production-basis (NOT consumption-basis MRIO)",
            "defra_kgco2e_per_gbp": sic21_intensity,
            "data_year": 2024,
            "mapping_confidence": "sourced-lower-bound",
            "factor_source": SOURCE,
            "mapping_rationale": RATIONALE,
        })
    return pd.DataFrame(rows)


def main():
    os.makedirs(INT_DIR, exist_ok=True)

    # ── Report DEFRA file contents (unchanged from v1) ─────────────────────
    print("=" * 70)
    print("DEFRA 2025 GHG FILE — SCOPE 3 CATEGORIES (operational, not spend-based)")
    print("=" * 70)
    wb = openpyxl.load_workbook(DEFRA_FLAT, read_only=True)
    ws = wb["Factors by Category"]
    rows = list(ws.iter_rows(values_only=True))
    seen = set()
    for r in rows[6:]:
        if r[1] and "3" in str(r[1]):
            key = (r[2], r[3])
            if key not in seen:
                seen.add(key)
                print(f"  {str(r[2]):<40} / {r[3]}")

    print()
    print("  -> ZERO spend-based (kgCO2e/GBP) factors in DEFRA 2025 file.")
    print("  -> All Scope 3 factors are physical-unit (per tonne, km, kWh).")

    # ── Show real ONS intensities ──────────────────────────────────────────
    print()
    print("=" * 70)
    print("ONS GHG INTENSITY BY INDUSTRY — REAL DATA (kgCO2e per GBP, 2024)")
    print("Source: ONS UK Environmental Accounts, 5 June 2026")
    print("=" * 70)
    ons_vals = get_ons_all_intensities(year=2024)
    for label, val in ons_vals.items():
        display = f"{val:.4f}" if val is not None else "SUPPRESSED (unreliable)"
        print(f"  {label:<60} {display}")

    sic21 = get_ons_sic21_intensity(year=2024)
    print()
    print(f"  SIC 21 (2024): {sic21} kgCO2e/GBP  <- PRIMARY FACTOR USED")
    print()
    print("  IMPORTANT: These are UK PRODUCTION-BASIS figures.")
    print("  The NHS Carbon Footprint model uses CONSUMPTION-BASIS MRIO factors")
    print("  (higher, because they include imports' embedded emissions).")
    print("  Those specific MRIO multipliers are NOT in any public download.")
    print("  The NHS Net Zero 2022 report (pp.56-58) confirms MRIO methodology")
    print("  but does not publish the multiplier table.")

    # ── Build and save mapping ─────────────────────────────────────────────
    mapping = build_mapping_df(sic21)

    print()
    print("=" * 70)
    print("REVISED BNF -> EMISSION FACTOR MAPPING")
    print("=" * 70)
    print(f"  Single factor applied to all 16 chapters: {sic21} kgCO2e/GBP")
    print(f"  Source: ONS SIC 21, UK production-basis, 2024")
    print(f"  Confidence: 'sourced-lower-bound'")
    print(f"  (No public data supports chapter-level differentiation)")
    print()
    for _, row in mapping.iterrows():
        print(f"  Ch {row['bnf_chapter_num']:>2} | {row['bnf_chapter_name']:<45} | {row['defra_kgco2e_per_gbp']:.4f}")

    out_path = os.path.join(INT_DIR, "bnf_to_defra_mapping.csv")
    mapping.to_csv(out_path, index=False)
    print(f"\nSaved: {out_path}")
    print()
    print("Next step: Run src/features/apply_carbon_labels.py to rebuild baseline.")


if __name__ == "__main__":
    main()
